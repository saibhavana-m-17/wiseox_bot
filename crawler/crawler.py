"""
WiseOx Knowledge Base Crawler
Reads URLs from an Excel file (organized by tabs), crawls accessible pages,
extracts textual content, and builds a knowledge_base/ directory with manifest.json.
Uses concurrent requests (20 workers) to handle 200+ links efficiently.
"""

import sys
import os
import re
import json
import requests
import openpyxl
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from collections import OrderedDict
from datetime import datetime, timezone

MAX_WORKERS = 20
TIMEOUT = 15
print_lock = Lock()


def load_urls_by_sheet(filepath):
    """Load URLs grouped by sheet name from an Excel file.

    Extracts both plain-text cell values and embedded hyperlinks.
    Deduplicates URLs across all tabs using a global seen set.

    Args:
        filepath: Path to the Excel file containing URLs.

    Returns:
        OrderedDict mapping sheet name to list of unique URLs.
    """
    wb = openpyxl.load_workbook(filepath)
    sheet_urls = OrderedDict()
    global_seen = set()

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        urls = []

        for row in ws.iter_rows():
            for cell in row:
                found_urls = []

                # Check for embedded hyperlink
                if cell.hyperlink and cell.hyperlink.target:
                    link = cell.hyperlink.target.strip()
                    if link.startswith("http"):
                        found_urls.append(link)

                # Check cell value as plain text URL
                val = cell.value
                if val and isinstance(val, str) and val.strip().startswith("http"):
                    found_urls.append(val.strip())

                for url in found_urls:
                    if url not in global_seen:
                        global_seen.add(url)
                        urls.append(url)

        if urls:
            sheet_urls[sheet] = urls

    return sheet_urls


def fetch_and_extract(url, timeout=TIMEOUT):
    """Fetch a URL and extract its main textual content.

    Strips HTML tags, scripts, styles, navigation, headers, and footers
    using BeautifulSoup. Returns extracted text on success or an error
    description on failure.

    Args:
        url: The URL to fetch.
        timeout: Request timeout in seconds.

    Returns:
        Tuple of (content, error) where content is the extracted text
        (or None on failure) and error is the error description (or None
        on success).
    """
    headers = {
        "User-Agent": "WiseOx-Crawler/1.0"
    }
    try:
        response = requests.get(url, timeout=timeout, headers=headers)
        if response.status_code != 200:
            return (None, f"HTTP {response.status_code} for {url}")

        soup = BeautifulSoup(response.text, "html.parser")

        # Remove non-content tags
        for tag in soup.find_all(["script", "style", "nav", "header", "footer"]):
            tag.decompose()

        text = soup.get_text(separator="\n")
        # Strip whitespace from each line and remove blank lines
        lines = [line.strip() for line in text.splitlines()]
        text = "\n".join(line for line in lines if line)

        return (text, None)

    except requests.exceptions.Timeout:
        return (None, f"Timeout after {timeout}s for {url}")
    except requests.exceptions.ConnectionError:
        return (None, f"Connection error for {url}")
    except requests.exceptions.RequestException as e:
        return (None, f"Request error for {url}: {e}")


def sanitize_filename(url):
    """Convert a URL to a safe filename for the knowledge base.

    Strips the protocol prefix (http:// or https://), replaces non-alphanumeric
    characters (except hyphens) with underscores, collapses consecutive underscores,
    strips leading/trailing underscores, and appends .txt.

    The transformation is deterministic and preserves enough URL structure to ensure
    distinct URLs produce distinct filenames.

    Args:
        url: The source URL.

    Returns:
        A sanitized filename string ending in .txt.
    """
    # Strip protocol prefix
    name = re.sub(r'^https?://', '', url)

    # Replace non-alphanumeric characters (except hyphens) with underscores
    name = re.sub(r'[^a-zA-Z0-9\-]', '_', name)

    # Collapse consecutive underscores
    name = re.sub(r'_+', '_', name)

    # Strip leading/trailing underscores
    name = name.strip('_')

    # Handle edge case where name is empty after stripping
    if not name:
        name = 'unnamed'

    return name + '.txt'


def truncate_content(content, max_chars=50000):
    """Truncate content that exceeds the maximum character limit.

    If content exceeds max_chars, truncates to max_chars characters and appends
    a truncation notice. Content within the limit is returned unchanged.

    Args:
        content: The text content to potentially truncate.
        max_chars: Maximum number of characters before truncation (default 50000).

    Returns:
        The original content if within limit, or truncated content with notice appended.
    """
    if len(content) <= max_chars:
        return content
    return content[:max_chars] + "\n\n[Content truncated at 50,000 characters]"


def build_knowledge_base(sheet_urls, output_dir="knowledge_base"):
    """Orchestrate concurrent fetching and build the knowledge base.

    Creates the output directory, fetches all URLs concurrently using
    ThreadPoolExecutor, saves extracted content as .txt files, and
    generates manifest.json with metadata for each entry.

    Args:
        sheet_urls: OrderedDict mapping sheet name to list of URLs.
        output_dir: Path to the output directory for content files.

    Returns:
        Dict containing the manifest data (entries, counts, timestamps).
    """
    os.makedirs(output_dir, exist_ok=True)

    # Build url -> sheet mapping
    url_to_sheet = {}
    all_urls = []
    for sheet, urls in sheet_urls.items():
        for url in urls:
            all_urls.append(url)
            url_to_sheet[url] = sheet

    total = len(all_urls)
    entries = []
    success_count = 0
    fail_count = 0

    print(f"\n⚡ Crawling {total} URLs with {MAX_WORKERS} parallel workers...\n")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(fetch_and_extract, url): (i, url)
            for i, url in enumerate(all_urls, 1)
        }
        for future in as_completed(futures):
            idx, url = futures[future]
            content, error = future.result()

            if error:
                with print_lock:
                    print(f"  [{idx}/{total}] ❌ {error}")
                fail_count += 1
            else:
                content = truncate_content(content)
                filename = sanitize_filename(url)
                filepath = os.path.join(output_dir, filename)
                try:
                    with open(filepath, "w", encoding="utf-8") as f:
                        f.write(content)
                except OSError as e:
                    with print_lock:
                        print(f"  [{idx}/{total}] ❌ Disk write error for {url}: {e}")
                    fail_count += 1
                    continue

                entries.append({
                    "url": url,
                    "filename": filename,
                    "sourceTab": url_to_sheet[url],
                    "crawledAt": datetime.now(timezone.utc).isoformat(),
                    "charCount": len(content),
                })
                success_count += 1
                with print_lock:
                    print(f"  [{idx}/{total}] ✅ {url[:90]}")

    manifest = {
        "crawledAt": datetime.now(timezone.utc).isoformat(),
        "totalUrls": total,
        "successCount": success_count,
        "failedCount": fail_count,
        "entries": entries,
    }

    manifest_path = os.path.join(output_dir, "manifest.json")
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2)

    return manifest


def main():
    """Entry point for the crawler.

    Loads URLs from links.xlsx, builds the knowledge base, and prints
    a crawl summary. Exits with status code 1 if the Excel file is
    missing or unreadable.
    """
    filepath = sys.argv[1] if len(sys.argv) > 1 else "links.xlsx"

    print(f"\n📂 Loading URLs from: {filepath}")
    try:
        sheet_urls = load_urls_by_sheet(filepath)
    except FileNotFoundError:
        print(f"❌ File not found: {filepath}")
        print("Usage: python3 crawler/crawler.py <path_to_excel_file>")
        sys.exit(1)

    if not sheet_urls:
        print("⚠️  No URLs found in the Excel file.")
        return

    for sheet, urls in sheet_urls.items():
        print(f"  📑 {sheet}: {len(urls)} links")

    total = sum(len(urls) for urls in sheet_urls.values())
    print(f"\n🔗 Total: {total} unique URLs across {len(sheet_urls)} tabs.")

    manifest = build_knowledge_base(sheet_urls)

    print("\n" + "=" * 70)
    print(f"  CRAWL COMPLETE: {manifest['successCount']} succeeded, "
          f"{manifest['failedCount']} failed (out of {manifest['totalUrls']})")
    print("=" * 70)
    print(f"\n📁 Knowledge base saved to: knowledge_base/")
    print(f"📄 Manifest: knowledge_base/manifest.json")


if __name__ == "__main__":
    main()
