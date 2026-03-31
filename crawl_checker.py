"""
Link Accessibility Checker
Reads URLs from an Excel file (organized by tabs) and checks accessibility.
Uses concurrent requests (20 workers) to handle 200+ links quickly.
Results are grouped by sheet/tab name.
"""

import sys
import requests
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from collections import OrderedDict

MAX_WORKERS = 20
TIMEOUT = 15
print_lock = Lock()


def load_urls_by_sheet(filepath):
    """Load URLs grouped by sheet name. Extracts both plain text URLs and embedded hyperlinks."""
    wb = openpyxl.load_workbook(filepath)
    sheet_urls = OrderedDict()
    global_seen = set()

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        urls = []

        for row in ws.iter_rows():
            for cell in row:
                found_urls = []

                # 1. Check for embedded hyperlink
                if cell.hyperlink and cell.hyperlink.target:
                    link = cell.hyperlink.target.strip()
                    if link.startswith("http"):
                        found_urls.append(link)

                # 2. Check cell value as plain text URL
                val = cell.value
                if val and isinstance(val, str) and val.strip().startswith("http"):
                    found_urls.append(val.strip())

                for url in found_urls:
                    if url not in global_seen:
                        global_seen.add(url)
                        urls.append(url)

        if urls:
            sheet_urls[sheet] = urls

    return sheet_urls


def check_url(index, total, url):
    """Check if a URL is accessible."""
    try:
        resp = requests.get(
            url, timeout=TIMEOUT,
            headers={"User-Agent": "Mozilla/5.0 (compatible; LinkChecker/1.0)"},
            allow_redirects=True,
        )
        ok, status, error = resp.ok, resp.status_code, None
    except requests.exceptions.Timeout:
        status, ok, error = None, False, "Timeout"
    except requests.exceptions.ConnectionError:
        status, ok, error = None, False, "Connection Error"
    except requests.exceptions.TooManyRedirects:
        status, ok, error = None, False, "Too Many Redirects"
    except Exception as e:
        status, ok, error = None, False, str(e)

    icon = "✅" if ok else "❌"
    detail = str(status) if ok else (f"HTTP {status}" if status else error)
    with print_lock:
        print(f"  [{index}/{total}] {icon} [{detail}] {url[:90]}")

    return url, status, ok, error


def main():
    filepath = sys.argv[1] if len(sys.argv) > 1 else "links.xlsx"

    print(f"\n📂 Loading URLs from: {filepath}")
    try:
        sheet_urls = load_urls_by_sheet(filepath)
    except FileNotFoundError:
        print(f"❌ File not found: {filepath}")
        print("Usage: python3 crawl_checker.py <path_to_excel_file>")
        sys.exit(1)

    if not sheet_urls:
        print("⚠️  No URLs found in the Excel file.")
        sys.exit(0)

    all_urls = []
    url_to_sheet = {}
    for sheet, urls in sheet_urls.items():
        print(f"  📑 {sheet}: {len(urls)} links")
        for u in urls:
            all_urls.append(u)
            url_to_sheet[u] = sheet

    total = len(all_urls)
    print(f"\n🔗 Total: {total} unique URLs across {len(sheet_urls)} tabs.")
    print(f"⚡ Checking with {MAX_WORKERS} parallel workers...\n")

    results = {}  # url -> (status, ok, error)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(check_url, i, total, url): url
            for i, url in enumerate(all_urls, 1)
        }
        for future in as_completed(futures):
            url, status, ok, error = future.result()
            results[url] = (status, ok, error)

    # Group results by sheet
    sheet_accessible = OrderedDict()
    sheet_not_accessible = OrderedDict()
    total_ok = 0
    total_fail = 0

    for sheet, urls in sheet_urls.items():
        sheet_accessible[sheet] = []
        sheet_not_accessible[sheet] = []
        for url in urls:
            status, ok, error = results[url]
            if ok:
                sheet_accessible[sheet].append((url, status))
                total_ok += 1
            else:
                reason = f"HTTP {status}" if status else error
                sheet_not_accessible[sheet].append((url, reason))
                total_fail += 1

    # Print summary
    print("\n" + "=" * 70)
    print(f"  RESULTS: {total_ok} accessible, {total_fail} not accessible (out of {total})")
    print("=" * 70)

    for sheet in sheet_urls:
        ok_list = sheet_accessible[sheet]
        fail_list = sheet_not_accessible[sheet]
        print(f"\n{'─' * 70}")
        print(f"  📑 TAB: {sheet}  ({len(ok_list)} ✅ / {len(fail_list)} ❌)")
        print(f"{'─' * 70}")

        if ok_list:
            print(f"\n  ✅ Accessible:")
            for url, status in sorted(ok_list):
                print(f"    [{status}] {url}")

        if fail_list:
            print(f"\n  ❌ Not Accessible:")
            for url, reason in sorted(fail_list):
                print(f"    [{reason}] {url}")

    # Save report
    report_path = "crawl_report.txt"
    with open(report_path, "w") as f:
        f.write("LINK ACCESSIBILITY REPORT\n")
        f.write(f"{'=' * 60}\n")
        f.write(f"Total URLs: {total}\n")
        f.write(f"Accessible: {total_ok}\n")
        f.write(f"Not Accessible: {total_fail}\n\n")

        for sheet in sheet_urls:
            ok_list = sheet_accessible[sheet]
            fail_list = sheet_not_accessible[sheet]
            f.write(f"{'─' * 60}\n")
            f.write(f"TAB: {sheet}  ({len(ok_list)} accessible / {len(fail_list)} not accessible)\n")
            f.write(f"{'─' * 60}\n")
            if ok_list:
                f.write("\n  ACCESSIBLE:\n")
                for url, status in sorted(ok_list):
                    f.write(f"    [{status}] {url}\n")
            if fail_list:
                f.write("\n  NOT ACCESSIBLE:\n")
                for url, reason in sorted(fail_list):
                    f.write(f"    [{reason}] {url}\n")
            f.write("\n")

    print(f"\n📄 Report saved to: {report_path}")


if __name__ == "__main__":
    main()
